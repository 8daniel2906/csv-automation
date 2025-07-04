import csv
import io
import json
import os

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import psycopg2 as psy
import requests
import sklearn
import tensorflow as tf
import uvicorn

from api import api_url_template
from datetime import datetime, timedelta, time
from fastapi import FastAPI
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from pydantic import BaseModel
from tensorflow.keras.losses import MeanSquaredError
from zoneinfo import ZoneInfo
import joblib
import plotly.graph_objects as go
import streamlit as st


def convert_json_row_to_arrays1(zeile_json: dict):
    zeile = zeile_json.copy()
    zeile["historic_prediction"] = np.array(zeile["historic_prediction"])
    zeile["blaue_kurve"] = np.array(zeile["blaue_kurve"])
    zeile["lower_hist"] = np.array(zeile["lower_hist"])
    zeile["upper_hist"] = np.array(zeile["upper_hist"])
    return zeile

def convert_json_row_to_arrays2(grouped_json: dict):
    converted = grouped_json.copy()

    converted["historic_prediction"] = [np.array(arr) for arr in grouped_json["historic_prediction"]]
    converted["historic_vergleich"] = [np.array(arr) for arr in grouped_json["historic_vergleich"]]
    converted["lower_hist"] = [np.array(arr) for arr in grouped_json["lower_hist"]]
    converted["upper_hist"] = [np.array(arr) for arr in grouped_json["upper_hist"]]
    converted["statistics"] = grouped_json["statistics"]

    return converted

def get_live_data():
    response = requests.get("https://image-api-latest-3.onrender.com/get-live")
    #response = requests.get("http://127.0.0.1:8000/get-live")
    response_json = response.json()
    converted = convert_json_row_to_arrays1(response_json["results"][0])
    return converted

def get_live_data2():
    response = requests.get("https://image-api-latest-3.onrender.com/get-live2")
    #response = requests.get("http://127.0.0.1:8000/get-live2")
    response_json = response.json()
    converted = convert_json_row_to_arrays2(response_json)
    return converted

def vorhersage_live_text(upper):
    vorhersage_text = ["<span style='color:red'>Hochwasserwarnung für die nächsten 12 Stunden!</span>",
                       "<span style='color:green'>Entwarnung für Hochwasser für die nächsten 12 Stunden!</span>"]
    return vorhersage_text[0] if np.max(upper) > hochwasser_schwelle else vorhersage_text[1]

def fast_now():
    now = datetime.now()
    fast_now = now.replace(minute=0, second=0, microsecond=0) - timedelta(hours=1)
    return fast_now.isoformat()
#macht die jetzige zeit auf minute bis sekunde 00:00 und 2 stunden zurück
def date_2_iso(jahr: int, monat: int, tag: int, stunde: int) -> str:
    dt = datetime(jahr, monat, tag, stunde, 0, 0)  # Minuten und Sekunden = 0
    return dt.isoformat()
def stunden_zurueck(iso_datum: str, k: int) -> str:
    dt = datetime.fromisoformat(iso_datum)
    neue_zeit = dt - timedelta(hours=k)
    return neue_zeit.isoformat()
def stunden_danach(iso_datum: str, k: int) -> str:
    dt = datetime.fromisoformat(iso_datum)
    neue_zeit = dt + timedelta(hours=k)
    return neue_zeit.isoformat()

def stunden_diff(iso_start: str, iso_ende: str) -> float:
    start = datetime.fromisoformat(iso_start)
    ende = datetime.fromisoformat(iso_ende)
    diff = ende - start
    return diff.total_seconds() / 3600  # Sekunden in Stunden umrechnen


def osw_api_extract(anfang: str, ende: str, api_url_template: str):
    anfang_minus_vierzehn_stunden = stunden_zurueck(anfang, 14)#für die vorhersagenberechnung
    api_url = api_url_template.format(start=anfang_minus_vierzehn_stunden, end=ende)
    response = requests.get(api_url)  # anfrage an die api wird in response gesaved
    json = response.json()
    return json


def json_to_dataframe(json_data, spalten_umbenennung=None):
    df = pd.DataFrame(json_data)
    if spalten_umbenennung:
        df.rename(columns=spalten_umbenennung, inplace=True)
    return df

def df_cleansing(df):
    df = df[df["Wert"] != 0]
    df["Zeit"] = pd.to_datetime(df["Zeit"])
    # Runde Startzeit auf Minute 0 und Endzeit auf Minute 59
    start_time = df["Zeit"].min().replace(minute=0, second=0, microsecond=0)
    end_time = df["Zeit"].max().replace(minute=59, second=0, microsecond=0)
    # Erstelle vollständige Zeitreihe in Minutenabstand
    df_full = pd.DataFrame({"Zeit": pd.date_range(start=start_time, end=end_time, freq="min")})
    # Merge & Interpolation
    df = df_full.merge(df, on="Zeit", how="left").interpolate(method="linear")
    df = df.fillna(method="bfill").fillna(method="ffill")  # optional: doppelte Absicherung

    # Zerlegen in Einzelkomponenten
    df[["Jahr", "Monat", "Tag", "Stunde", "Minute"]] = df["Zeit"].apply(
        lambda x: [x.year, x.month, x.day, x.hour, x.minute]
    ).apply(pd.Series)
    df = df.drop(columns=["Zeit"])[["Jahr", "Monat", "Tag", "Stunde", "Minute", "Wert"]]
    if df.shape[0] % 60 != 0:
        print("Fehler: Anzahl der Zeilen ist nicht teilbar durch 60")
    return df


def df_feature_engineering(df):
    df = df.copy()
    # Jahr - 200 und durch 1000 teilen
    df['Jahr'] = (df['Jahr'] - 2000) / 1000.0

    # Monat zyklisch in Sinus und Kosinus umwandeln
    df['Monat_sin'] = np.sin(2 * np.pi * df['Monat'] / 12)
    df['Monat_cos'] = np.cos(2 * np.pi * df['Monat'] / 12)

    # Tag zyklisch in Sinus und Kosinus umwandeln
    df['Tag_sin'] = np.sin(2 * np.pi * df['Tag'] / 31)
    df['Tag_cos'] = np.cos(2 * np.pi * df['Tag'] / 31)

    # Stunde zyklisch in Sinus und Kosinus umwandeln
    df['Stunde_sin'] = np.sin(2 * np.pi * df['Stunde'] / 24)
    df['Stunde_cos'] = np.cos(2 * np.pi * df['Stunde'] / 24)

    # Minute zyklisch in Sinus und Kosinus umwandeln
    df['Minute_sin'] = np.sin(2 * np.pi * df['Minute'] / 60)
    df['Minute_cos'] = np.cos(2 * np.pi * df['Minute'] / 60)

    # Wert durch 1000 teilen
    df['Wert'] = df['Wert'] / 1000.0

    # Die relevanten Spalten in der gewünschten Reihenfolge
    df = df[['Jahr', 'Monat_sin', 'Monat_cos', 'Tag_sin', 'Tag_cos',
             'Stunde_sin', 'Stunde_cos', 'Minute_sin', 'Minute_cos', 'Wert']]

    df.columns = ["col1", "col2", "col3", "col4", "col5", "col6", "col7", "col8", "col9", "water_level"]
    # mehr features , diesmal moving averages #minutes_per_week und minutes_per_month sind hier aber erstmal gleich inhaltlich ,weil ich nur daten der letzen woche entnehme
    minutes_per_day = 60 * 24  # tägliche Durchschnitte
    minutes_per_week = minutes_per_day * 7  # wöchentliche durchschnitte
    minutes_per_month = minutes_per_day * 30  # monatliche Durchschnitte !!!! das feature macht eigentlich noch keinen sinn , wenn ich nur für eine Woche  die Vorhersagendemonstration mache
    # Berechnung der Moving Averages
    df["MA_Day"] = df["water_level"].rolling(window=minutes_per_day, min_periods=1).mean()
    df["MA_Week"] = df["water_level"].rolling(window=minutes_per_week, min_periods=1).mean()
    df["MA_Month"] = df["water_level"].rolling(window=minutes_per_month, min_periods=1).mean()
    # Spalten neu anordnen (Moving Averages an den Anfang)
    df_transformed = df[["MA_Month", "MA_Week", "MA_Day"] + df.columns[:-3].tolist()]

    # Numpy-Array erstellen
    return df_transformed.to_numpy()

def get_models():
    model = tf.keras.models.load_model("fnn.h5",custom_objects={"mse": MeanSquaredError()})
    ####
    model_bundle = joblib.load('models/kmeans_10cl_10timebin_180chunk_90%e_80%PI_750schwelle.pkl')
    kmeans = model_bundle['kmeans']
    interval_matrix = model_bundle['interval_matrix']
    chunk = 180
    cluster = 10
    time_teile = 10
    time_labels_klein = assign_time_bins_full_array(720, time_teile, 720)
    return model, kmeans, interval_matrix, chunk, cluster, time_teile, time_labels_klein

def assign_clusters_inference(prediction, chunk_size, trained_kmeans):
    chunks = []
    chunk_ranges = []
    for start in range(0, len(prediction), chunk_size):
        end = start + chunk_size
        if end <= len(prediction):
            chunks.append(prediction[start:end])
            chunk_ranges.append((start, end))
    chunks = np.array(chunks)
    predicted_labels = trained_kmeans.predict(chunks)
    cluster_labels = np.full(len(prediction), np.nan)
    for i, (start, end) in enumerate(chunk_ranges):
        cluster_labels[start:end] = predicted_labels[i]
    return cluster_labels

def assign_time_bins_full_array(array_len, n_bins, packet_size):
    import numpy as np

    assert array_len % packet_size == 0, f"Länge muss ein Vielfaches von {packet_size} sein!"

    full_time_bin_array = []

    for packet_start in range(0, array_len, packet_size):
        time_bins = np.linspace(0, packet_size, n_bins + 1)
        # Lokale Indizes innerhalb des Pakets (0–719)
        local_indices = np.arange(packet_size)
        local_bins = np.digitize(local_indices, time_bins, right=False) - 1
        full_time_bin_array.extend(local_bins)

    return np.array(full_time_bin_array)

def get_quantile_bounds_from_labels(cluster_labels, time_labels, interval_matrix):
    assert len(cluster_labels) == len(time_labels), "Label-Arrays müssen gleich lang sein."

    lo_array = np.full(len(cluster_labels), np.nan)
    hi_array = np.full(len(cluster_labels), np.nan)

    for i in range(len(cluster_labels)):
        c = int(cluster_labels[i])
        t = int(time_labels[i])
        if not (np.isnan(c) or np.isnan(t)):
            lo, hi = interval_matrix[c][t]
            lo_array[i] = lo
            hi_array[i] = hi

    return lo_array, hi_array

def inference(steps, array, start):
    #temp_arr = []
    #historic_prediction = []
    #historic_prediction_temp = []
    results = []
    model, kmeans, interval_matrix, chunk, cluster, time_teile, time_labels_klein = get_models()


    for i in range(steps):
        historic_datum = array[ i * 60, :12].reshape(12, )
        historic_stand = array[ i * 60 : ( i * 60) + 840, 12].reshape(840, )
        historic_vergleich = array[( i * 60) + 840 : ( i * 60) + 840 + 720, 12]*1000
        temp_arr = np.concatenate((historic_datum, historic_stand))
        historic_prediction = []
        for j in range(2):
            historic_prediction_temp = model.predict(temp_arr.reshape(1, -1), verbose=0).reshape(360, )
            historic_prediction = np.append(historic_prediction, historic_prediction_temp)
            temp_arr = temp_arr[12:]
            temp_arr = np.concatenate((temp_arr, historic_prediction_temp))
            temp_arr = temp_arr[360:]
            temp_arr = np.concatenate((array[ i * 60 + (j + 1) * 360, :12].reshape(12, ), temp_arr))

        historic_prediction = historic_prediction * 1000
        cluster_labels2 = assign_clusters_inference( historic_prediction, chunk, kmeans)  # chunk, kmeans global
        q_low_hist, q_up_hist = get_quantile_bounds_from_labels(cluster_labels2, time_labels_klein, interval_matrix)
        lower_hist = historic_prediction - np.abs(q_low_hist)
        upper_hist = historic_prediction + np.abs(q_up_hist)
        zeit1 = start
        zeit2 = stunden_danach(start, 12)
        start = stunden_danach(start, 1)
        zeile = [zeit1, zeit2, np.max(np.array(historic_vergleich)), np.max(np.array(upper_hist)), np.array(historic_prediction), np.array(historic_vergleich), np.array(lower_hist), np.array(upper_hist)]
        results.append(zeile)

    return results


def extract_and_tranform(zeitpunkt1, zeitpunkt2, api_template ):

    db_ende =  zeitpunkt1
    db_start = stunden_zurueck(db_ende, 12)
    json_daten = osw_api_extract(stunden_zurueck(db_ende, 25), zeitpunkt2, api_url_template)
    df = json_to_dataframe(json_daten, spalten_umbenennung={"begin": "Zeit", "v": "Wert"})
    df2 = df_cleansing(df)
    df3 = df_feature_engineering(df2)
    steps = int(stunden_diff(db_ende, zeitpunkt2))
    results = inference(steps, df3, stunden_danach(db_start, 1))

    return results

def get_latest_endzeitpunkt_iso(conn_str):
    query = "SELECT MAX(endzeit) FROM zeitreihe_metadata;"
    try:
        with psy.connect(conn_str) as conn:
            with conn.cursor() as cur:
                cur.execute(query)
                result = cur.fetchone()[0]
                if result:
                    # In ISO-Format umwandeln: 2025-03-06T00:00:00
                    return result.strftime("%Y-%m-%dT%H:%M:%S")
                else:
                    return None  # Falls Tabelle leer
    except Exception as e:
        print(f"Fehler: {e}")
        return None

def load_in_db(conn_str, results):
    with psy.connect(conn_str) as conn:
        with conn.cursor() as cur:
            for zeile in results:
                zeit1 = zeile[0]
                zeit2 = zeile[1]
                max_vergleich = zeile[2]
                max_upper = zeile[3]
                prediction = zeile[4]
                vergleich = zeile[5]
                lower = zeile[6]
                upper = zeile[7]

                # Prüfen, ob Kombination bereits vorhanden ist
                cur.execute("""
                    SELECT id FROM zeitreihe_metadata
                    WHERE startzeit = %s AND endzeit = %s
                """, (zeit1, zeit2))
                exists = cur.fetchone()

                if exists:
                    print(f"⚠️ Kombination {zeit1} - {zeit2} existiert bereits. Skip.")
                    continue

                # Insert in metadata
                cur.execute("""
                    INSERT INTO zeitreihe_metadata (startzeit, endzeit, max_value_historic, max_value_upperpi_80_perc)
                    VALUES (%s, %s, %s, %s)
                    RETURNING id
                """, (zeit1, zeit2, round(max_vergleich, 2), round(max_upper, 2)))
                metadata_id = cur.fetchone()[0]

                datenpunkte = []
                for minute in range(720):
                    wert_pred = round(prediction[minute], 2)
                    wert_vergl = round(vergleich[minute], 2)
                    wert_lower = round(lower[minute], 2)
                    wert_upper = round(upper[minute], 2)
                    datenpunkte.append((metadata_id, minute, wert_pred, wert_vergl, wert_lower, wert_upper))

                # Bulk insert Zeitreihen
                cur.executemany("""
                    INSERT INTO zeitreihe_daten
                    (metadata_id, minute_value, prediction_value, historic_value, lowerpi_value, upperpi_value)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, datenpunkte)
                print("done")

            conn.commit()




def load_in_db2(conn_str, results):
    with psy.connect(conn_str) as conn:
        with conn.cursor() as cur:
            # Vorhandene Kombinationen laden
            cur.execute("SELECT startzeit, endzeit FROM zeitreihe_metadata")
            vorhandene = set(cur.fetchall())

            buffer = io.StringIO()

            for zeile in results:
                zeit1, zeit2, max_vergleich, max_upper, prediction, vergleich, lower, upper = zeile

                if (zeit1, zeit2) in vorhandene:
                    print(f"⚠️ Kombination {zeit1} - {zeit2} existiert bereits. Skip.")
                    continue

                # Metadata insert
                cur.execute("""
                    INSERT INTO zeitreihe_metadata (startzeit, endzeit, max_value_historic, max_value_upperpi_80_perc)
                    VALUES (%s, %s, %s, %s)
                    RETURNING id
                """, (zeit1, zeit2, int(max_vergleich), int(max_upper)))
                metadata_id = cur.fetchone()[0]

                for i in range(144):
                    buffer.write(f"{metadata_id}\t{i}\t{round(prediction[i*5], 1)}\t{round(vergleich[i*5], 1)}\t{round(lower[i*5], 1)}\t{round(upper[i*5], 1)}\n")

                print(f"✅ Metadata ID {metadata_id} vorbereitet.")

            buffer.seek(0)

            cur.copy_from(
                buffer,
                'zeitreihe_daten',
                sep='\t',
                columns=('metadata_id', 'minute_value', 'prediction_value', 'historic_value', 'lowerpi_value', 'upperpi_value')
            )

            conn.commit()

            print("✅ Alle Daten mit COPY (psycopg2) eingespielt.")


###############################################################################

api_url_template = "https://api.opensensorweb.de/v1/organizations/open/networks/BAFG/devices/5952025/sensors/W/measurements/raw?start={start}%2B02:00&end={end}%2B02:00&interpolator=LINEAR"
conn_str = "postgresql://neondb_owner:npg_mPqZi9CG2txF@ep-divine-mud-a90zxdvg-pooler.gwc.azure.neon.tech/neondb?sslmode=require&channel_binding=require"

def warnung_generator(upper_peak, hist_peak):
    schwelle = 750 #das ist PI-model abhängig und kmeans abhängig
    hist_peak = np.array(hist_peak)
    upper_peak = np.array(upper_peak)
    mask11 = hist_peak >= schwelle      # Ground Truth: Überschwemmung
    mask12 = hist_peak < schwelle       # Ground Truth: keine Überschwemmung

    mask21 = upper_peak >= schwelle     # Prediction: Warnung
    mask22 = upper_peak < schwelle      # Prediction: keine Warnung

        # TP, FN, FP, TN
    TP = np.sum(mask11 & mask21)
    FN = np.sum(mask11 & mask22)
    FP = np.sum(mask12 & mask21)
    TN = np.sum(mask12 & mask22)

    #print(f"Warn.und HW.: {TP}, Entwarn. aber HW: {FN}, Warn. aber K.HW: {FP}, Entwarn. und K. HW: {TN} ")
    return int(TP), int(FN), int(FP), int(TN)


def fast_now():
    now = datetime.now()
    fast_now = now.replace(minute=0, second=0, microsecond=0) - timedelta(hours=1)
    return fast_now.isoformat()
#macht die jetzige zeit auf minute bis sekunde 00:00 und 2 stunden zurück
def date_2_iso(jahr: int, monat: int, tag: int, stunde: int) -> str:
    dt = datetime(jahr, monat, tag, stunde, 0, 0)  # Minuten und Sekunden = 0
    return dt.isoformat()
def stunden_zurueck(iso_datum: str, k: int) -> str:
    dt = datetime.fromisoformat(iso_datum)
    neue_zeit = dt - timedelta(hours=k)
    return neue_zeit.isoformat()
def stunden_danach(iso_datum: str, k: int) -> str:
    dt = datetime.fromisoformat(iso_datum)
    neue_zeit = dt + timedelta(hours=k)
    return neue_zeit.isoformat()

def stunden_diff(iso_start: str, iso_ende: str) -> float:
    start = datetime.fromisoformat(iso_start)
    ende = datetime.fromisoformat(iso_ende)
    diff = ende - start
    return diff.total_seconds() / 3600  # Sekunden in Stunden umrechnen


def osw_api_extract(anfang: str, ende: str, api_url_template: str):
    anfang_minus_vierzehn_stunden = stunden_zurueck(anfang, 14)#für die vorhersagenberechnung
    api_url = api_url_template.format(start=anfang_minus_vierzehn_stunden, end=ende)
    response = requests.get(api_url)  # anfrage an die api wird in response gesaved
    json = response.json()
    return json


def json_to_dataframe(json_data, spalten_umbenennung=None):
    df = pd.DataFrame(json_data)
    if spalten_umbenennung:
        df.rename(columns=spalten_umbenennung, inplace=True)
    return df

def df_cleansing(df):
    df = df[df["Wert"] != 0]
    df["Zeit"] = pd.to_datetime(df["Zeit"])
    # Runde Startzeit auf Minute 0 und Endzeit auf Minute 59
    start_time = df["Zeit"].min().replace(minute=0, second=0, microsecond=0)
    end_time = df["Zeit"].max().replace(minute=59, second=0, microsecond=0)
    # Erstelle vollständige Zeitreihe in Minutenabstand
    df_full = pd.DataFrame({"Zeit": pd.date_range(start=start_time, end=end_time, freq="min")})
    # Merge & Interpolation
    df = df_full.merge(df, on="Zeit", how="left").interpolate(method="linear")
    df = df.fillna(method="bfill").fillna(method="ffill")  # optional: doppelte Absicherung

    # Zerlegen in Einzelkomponenten
    df[["Jahr", "Monat", "Tag", "Stunde", "Minute"]] = df["Zeit"].apply(
        lambda x: [x.year, x.month, x.day, x.hour, x.minute]
    ).apply(pd.Series)
    df = df.drop(columns=["Zeit"])[["Jahr", "Monat", "Tag", "Stunde", "Minute", "Wert"]]
    if df.shape[0] % 60 != 0:
        print("Fehler: Anzahl der Zeilen ist nicht teilbar durch 60")
    return df


def df_feature_engineering(df):
    df = df.copy()
    # Jahr - 200 und durch 1000 teilen
    df['Jahr'] = (df['Jahr'] - 2000) / 1000.0

    # Monat zyklisch in Sinus und Kosinus umwandeln
    df['Monat_sin'] = np.sin(2 * np.pi * df['Monat'] / 12)
    df['Monat_cos'] = np.cos(2 * np.pi * df['Monat'] / 12)

    # Tag zyklisch in Sinus und Kosinus umwandeln
    df['Tag_sin'] = np.sin(2 * np.pi * df['Tag'] / 31)
    df['Tag_cos'] = np.cos(2 * np.pi * df['Tag'] / 31)

    # Stunde zyklisch in Sinus und Kosinus umwandeln
    df['Stunde_sin'] = np.sin(2 * np.pi * df['Stunde'] / 24)
    df['Stunde_cos'] = np.cos(2 * np.pi * df['Stunde'] / 24)

    # Minute zyklisch in Sinus und Kosinus umwandeln
    df['Minute_sin'] = np.sin(2 * np.pi * df['Minute'] / 60)
    df['Minute_cos'] = np.cos(2 * np.pi * df['Minute'] / 60)

    # Wert durch 1000 teilen
    df['Wert'] = df['Wert'] / 1000.0

    # Die relevanten Spalten in der gewünschten Reihenfolge
    df = df[['Jahr', 'Monat_sin', 'Monat_cos', 'Tag_sin', 'Tag_cos',
             'Stunde_sin', 'Stunde_cos', 'Minute_sin', 'Minute_cos', 'Wert']]

    df.columns = ["col1", "col2", "col3", "col4", "col5", "col6", "col7", "col8", "col9", "water_level"]
    # mehr features , diesmal moving averages #minutes_per_week und minutes_per_month sind hier aber erstmal gleich inhaltlich ,weil ich nur daten der letzen woche entnehme
    minutes_per_day = 60 * 24  # tägliche Durchschnitte
    minutes_per_week = minutes_per_day * 7  # wöchentliche durchschnitte
    minutes_per_month = minutes_per_day * 30  # monatliche Durchschnitte !!!! das feature macht eigentlich noch keinen sinn , wenn ich nur für eine Woche  die Vorhersagendemonstration mache
    # Berechnung der Moving Averages
    df["MA_Day"] = df["water_level"].rolling(window=minutes_per_day, min_periods=1).mean()
    df["MA_Week"] = df["water_level"].rolling(window=minutes_per_week, min_periods=1).mean()
    df["MA_Month"] = df["water_level"].rolling(window=minutes_per_month, min_periods=1).mean()
    # Spalten neu anordnen (Moving Averages an den Anfang)
    df_transformed = df[["MA_Month", "MA_Week", "MA_Day"] + df.columns[:-3].tolist()]

    # Numpy-Array erstellen
    return df_transformed.to_numpy()

def get_models():
    model = tf.keras.models.load_model("fnn.h5",custom_objects={"mse": MeanSquaredError()})
    ####
    model_bundle = joblib.load('models/kmeans_10cl_10timebin_180chunk_90%e_80%PI_750schwelle.pkl')
    kmeans = model_bundle['kmeans']
    interval_matrix = model_bundle['interval_matrix']
    chunk = 180
    cluster = 10
    time_teile = 10
    time_labels_klein = assign_time_bins_full_array(720, time_teile, 720)
    return model, kmeans, interval_matrix, chunk, cluster, time_teile, time_labels_klein

def assign_clusters_inference(prediction, chunk_size, trained_kmeans):
    chunks = []
    chunk_ranges = []
    for start in range(0, len(prediction), chunk_size):
        end = start + chunk_size
        if end <= len(prediction):
            chunks.append(prediction[start:end])
            chunk_ranges.append((start, end))
    chunks = np.array(chunks)
    predicted_labels = trained_kmeans.predict(chunks)
    cluster_labels = np.full(len(prediction), np.nan)
    for i, (start, end) in enumerate(chunk_ranges):
        cluster_labels[start:end] = predicted_labels[i]
    return cluster_labels

def assign_time_bins_full_array(array_len, n_bins, packet_size):
    import numpy as np

    assert array_len % packet_size == 0, f"Länge muss ein Vielfaches von {packet_size} sein!"

    full_time_bin_array = []

    for packet_start in range(0, array_len, packet_size):
        time_bins = np.linspace(0, packet_size, n_bins + 1)
        # Lokale Indizes innerhalb des Pakets (0–719)
        local_indices = np.arange(packet_size)
        local_bins = np.digitize(local_indices, time_bins, right=False) - 1
        full_time_bin_array.extend(local_bins)

    return np.array(full_time_bin_array)

def get_quantile_bounds_from_labels(cluster_labels, time_labels, interval_matrix):
    assert len(cluster_labels) == len(time_labels), "Label-Arrays müssen gleich lang sein."

    lo_array = np.full(len(cluster_labels), np.nan)
    hi_array = np.full(len(cluster_labels), np.nan)

    for i in range(len(cluster_labels)):
        c = int(cluster_labels[i])
        t = int(time_labels[i])
        if not (np.isnan(c) or np.isnan(t)):
            lo, hi = interval_matrix[c][t]
            lo_array[i] = lo
            hi_array[i] = hi

    return lo_array, hi_array

def inference_live(array, start):
    #temp_arr = []
    #historic_prediction = []
    #historic_prediction_temp = []
    results = []
    model, kmeans, interval_matrix, chunk, cluster, time_teile, time_labels_klein = get_models()
    for i in range(1):
        historic_datum = array[ i * 60, :12].reshape(12, )
        historic_stand = array[ i * 60 : ( i * 60) + 840, 12].reshape(840, )
        historic_vergleich = array[( i * 60) + 840 : ( i * 60) + 840 + 720, 12]*1000
        temp_arr = np.concatenate((historic_datum, historic_stand))
        blaue_kurve = historic_stand
        historic_prediction = []
        for j in range(2):
            historic_prediction_temp = model.predict(temp_arr.reshape(1, -1), verbose=0).reshape(360, )
            historic_prediction = np.append(historic_prediction, historic_prediction_temp)
            temp_arr = temp_arr[12:]
            temp_arr = np.concatenate((temp_arr, historic_prediction_temp))
            temp_arr = temp_arr[360:]
            temp_arr = np.concatenate((array[ i * 60 + (j + 1) * 360, :12].reshape(12, ), temp_arr))
        historic_prediction = historic_prediction * 1000
        cluster_labels2 = assign_clusters_inference( historic_prediction, chunk, kmeans)  # chunk, kmeans global
        q_low_hist, q_up_hist = get_quantile_bounds_from_labels(cluster_labels2, time_labels_klein, interval_matrix)
        lower_hist = historic_prediction - np.abs(q_low_hist)
        upper_hist = historic_prediction + np.abs(q_up_hist)
        zeit2 = stunden_danach(start, 24)
        zeit1 = stunden_zurueck(start, 2)
        zeile = [zeit1, zeit2, np.max(np.array(historic_vergleich)), np.max(np.array(upper_hist)), np.array(historic_prediction), np.array(blaue_kurve)*1000, np.array(lower_hist), np.array(upper_hist)]
        results.append(zeile)
    return results

def inference_live2( array, start):
    #temp_arr = []
    #historic_prediction = []
    #historic_prediction_temp = []
    results = []
    steps = len(array)//720 -1
    model, kmeans, interval_matrix, chunk, cluster, time_teile, time_labels_klein = get_models()
    for i in range(steps):
        historic_datum = array[ i * 720, :12].reshape(12, )
        historic_stand = array[ i * 720 : ( i * 720) + 840, 12].reshape(840, )
        historic_vergleich = array[( i * 720) + 840 : ( i * 720) + 840 + 720, 12]*1000
        temp_arr = np.concatenate((historic_datum, historic_stand))
        historic_prediction = []
        for j in range(2):
            historic_prediction_temp = model.predict(temp_arr.reshape(1, -1), verbose=0).reshape(360, )
            historic_prediction = np.append(historic_prediction, historic_prediction_temp)
            temp_arr = temp_arr[12:]
            temp_arr = np.concatenate((temp_arr, historic_prediction_temp))
            temp_arr = temp_arr[360:]
            temp_arr = np.concatenate((array[ i * 720 + (j + 1) * 360, :12].reshape(12, ), temp_arr))
        historic_prediction = historic_prediction * 1000
        cluster_labels2 = assign_clusters_inference( historic_prediction, chunk, kmeans)  # chunk, kmeans global
        q_low_hist, q_up_hist = get_quantile_bounds_from_labels(cluster_labels2, time_labels_klein, interval_matrix)
        lower_hist = historic_prediction - np.abs(q_low_hist)
        upper_hist = historic_prediction + np.abs(q_up_hist)
        zeit1 = start
        zeit2 = stunden_danach(start, 12)
        start = stunden_danach(start, 1)
        zeile = [zeit1, zeit2, np.max(np.array(historic_vergleich)), np.max(np.array(upper_hist)), np.array(historic_prediction), np.array(historic_vergleich), np.array(lower_hist), np.array(upper_hist)]
        results.append(zeile)
    return results

def extract_and_transform_live(stunden):
    jetzt = datetime.now(ZoneInfo("Europe/Berlin")).replace(tzinfo=None)  # in format beispielsweise: 2025-03-06T00:00:00
    jetzt = jetzt.isoformat()

    zeit1 =  stunden_zurueck(jetzt, stunden)
    json_daten = osw_api_extract(zeit1, jetzt, api_url_template)
    df = json_to_dataframe(json_daten, spalten_umbenennung={"begin": "Zeit", "v": "Wert"})
    df2 = df_cleansing(df)
    df3 = df_feature_engineering(df2)
    results = inference_live( df3, stunden_zurueck(jetzt, stunden))
    return results

def extract_and_transform_live2(stunden):
    jetzt = datetime.now(ZoneInfo("Europe/Berlin")).replace(tzinfo=None)  # in format beispielsweise: 2025-03-06T00:00:00
    jetzt = jetzt.isoformat()
    zeit1 =  stunden_zurueck(jetzt, stunden)
    json_daten = osw_api_extract(zeit1, jetzt, api_url_template)
    df = json_to_dataframe(json_daten, spalten_umbenennung={"begin": "Zeit", "v": "Wert"})
    df2 = df_cleansing(df)
    df3 = df_feature_engineering(df2)
    results = inference_live2( df3, zeit1)
    return results

def get_latest_endzeitpunkt_iso(conn_str):
    query = "SELECT MAX(endzeit) FROM zeitreihe_metadata;"
    try:
        with psy.connect(conn_str) as conn:
            with conn.cursor() as cur:
                cur.execute(query)
                result = cur.fetchone()[0]
                if result:
                    # In ISO-Format umwandeln: 2025-03-06T00:00:00
                    return result.strftime("%Y-%m-%dT%H:%M:%S")
                else:
                    return None  # Falls Tabelle leer
    except Exception as e:
        print(f"Fehler: {e}")
        return None
#########################################################################

def load_time_series(conn_str, startzeit_iso, endzeit_iso):
    with psy.connect(conn_str) as conn:
        with conn.cursor() as cur:
            current_start = startzeit_iso
            results = []

            while True:
                # Hole den ersten Datensatz, dessen startzeit = current_start ist
                cur.execute("""
                    SELECT id, startzeit, endzeit
                    FROM zeitreihe_metadata
                    WHERE startzeit = %s
                      AND endzeit <= %s
                    ORDER BY endzeit ASC
                    LIMIT 1
                """, (current_start, endzeit_iso))
                row = cur.fetchone()

                if not row:
                    print(f"⚠️ Keine weiteren Zeiträume gefunden ab {current_start}. Abbruch.")
                    break

                metadata_id, startzeit, endzeit = row

                # Hole Zeitreihendaten für dieses metadata_id
                cur.execute("""
                    SELECT minute_value, prediction_value, historic_value, lowerpi_value, upperpi_value
                    FROM zeitreihe_daten
                    WHERE metadata_id = %s
                    ORDER BY minute_value ASC
                """, (metadata_id,))
                data_rows = cur.fetchall()

                # Check, ob wir weniger als 720 Minuten haben
                num_points = len(data_rows)
                if num_points < 144:
                    print(f"⚠️ Zeitreihe bei {startzeit} - {endzeit} hat nur {num_points} Punkte (truncated).")
                elif num_points > 144:
                    print(f"⚠️ Zeitreihe länger als 720 Punkte, wird abgeschnitten.")
                    data_rows = data_rows[:144]

                # Speichere Ergebnis
                results.append({
                    "metadata_id": metadata_id,
                    "startzeit": startzeit,
                    "endzeit": endzeit,
                    "data": data_rows
                })

                # Nächster Startzeitpunkt = aktueller Endzeitpunkt
                current_start = endzeit.strftime("%Y-%m-%dT%H:%M:%S")

                # Prüfen, ob wir über Endzeit hinaus sind
                if current_start >= endzeit_iso:
                    print("✅ Endzeit erreicht oder überschritten. Fertig.")
                    break

            return results



def stretch_array(arr):
    arr = np.array(arr)
    n = len(arr)

    if n < 2:
        raise ValueError("Array muss mindestens 2 Elemente haben, um zu interpolieren.")

    # Für jedes Paar von Punkten wollen wir 5 Punkte insgesamt (Originalpunkte + 4 dazwischen)
    # Z.B. [a, b] -> [a, ..., b] mit 5 Punkten zwischen a und b
    # Wir bauen einen Index-Vektor von 0 bis n-1, jeweils in 1er-Schritten
    old_indices = np.arange(n)

    # Neuer Index-Vektor mit feinerer Auflösung
    new_indices = np.linspace(0, n - 1, (n - 1) * 5 + 1)

    # Interpolation
    new_arr = np.interp(new_indices, old_indices, arr)

    return new_arr.tolist()

def extract_and_stretch(results):
    """Extrahiert Spalten aus einem Block und interpoliert sie."""
    pred_ = []
    hist_ = []
    lower_ = []
    upper_ = []
    for i in range(len(results)):
        pred  = stretch_array([row[1] for row in results[i]['data']])
        hist  = stretch_array([row[2] for row in results[i]['data']])
        lower = stretch_array([row[3] for row in results[i]['data']])
        upper = stretch_array([row[4] for row in results[i]['data']])
        pred_.append(pred)
        hist_.append(hist)
        lower_.append(lower)
        upper_.append(upper)
    return np.array(pred_).flatten(), np.array(hist_).flatten(), np.array(lower_).flatten(), np.array(upper_).flatten()

def plot_time_series(pred, hist, lower, upper):
    """Plottet die Zeitreihen."""

    plt.figure(figsize=(14, 6))
    plt.plot(pred, label="Prediction")
    plt.plot(hist, label="Historic")
    plt.plot(lower, label="Lower PI", linestyle="--")
    plt.plot(upper, label="Upper PI", linestyle="--")
    plt.fill_between(range(len(pred)), lower, upper, alpha=0.2, color='gray')
    plt.legend()
    plt.title("Interpolierte Zeitreihen")
    plt.grid(True)
    plt.show()


def plot_time_series(pred, hist, lower, upper, save_path=None):
    plt.figure(figsize=(14, 6))
    plt.plot(pred, label="Prediction")
    plt.plot(hist, label="Historic")
    plt.plot(lower, label="Lower PI", linestyle="--")
    plt.plot(upper, label="Upper PI", linestyle="--")
    plt.fill_between(range(len(pred)), lower, upper, alpha=0.2, color='gray')
    plt.legend()
    plt.title("Interpolierte Zeitreihen")
    plt.grid(True)
    if save_path:
        plt.savefig(save_path)
        print(f"📈 Plot gespeichert unter {save_path}")
    plt.show()


def save_to_excel(pred, hist, lower, upper, filename):
    df = pd.DataFrame({
        "Prediction": pred,
        "Historic": hist,
        "Lower_PI": lower,
        "Upper_PI": upper
    })
    df.index.name = "Index"
    df.to_excel(filename)
    print(f"📄 Daten gespeichert in {filename}")

def export_to_excel_with_chart(pred, hist, lower, upper, filename="zeitreihe.xlsx"):
    # Baue DataFrame
    df = pd.DataFrame({
        "Prediction": pred,
        "Historic": hist,
        "Lower_PI": lower,
        "Upper_PI": upper
    })

    # Speichere erstmal ohne Chart
    df.to_excel(filename, index=False)

    # Lade Workbook mit openpyxl
    wb = load_workbook(filename)
    ws = wb.active

    # Erstelle LineChart
    chart = LineChart()
    chart.title = "Zeitreihe mit Prediction, Historic & PI"
    chart.y_axis.title = "Wert"
    chart.x_axis.title = "Zeitpunkte"

    # Referenz: alle Spalten (A bis D), Zeilen 2 bis letzte
    data = Reference(ws, min_col=1, max_col=4, min_row=1, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)

    # Positioniere Chart (z. B. ab Zelle F2)
    ws.add_chart(chart, "F2")

    # Speichern
    wb.save(filename)
    print(f"✅ Excel-Datei mit Diagramm gespeichert: {filename}")
    return filename
