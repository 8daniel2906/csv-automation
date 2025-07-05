from fastapi import FastAPI
from fastapi.responses import StreamingResponse
import io
import uvicorn
import os
from pydantic import BaseModel
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from fastapi.responses import JSONResponse
import matplotlib.pyplot as plt
from config import api_url_template, conn_str
from utils import *
###############################################################################
conn_str = "postgresql://neondb_owner:npg_mPqZi9CG2txF@ep-divine-mud-a90zxdvg-pooler.gwc.azure.neon.tech/neondb?sslmode=require&channel_binding=require"

def warnung_generator(upper_peak, hist_peak):
    schwelle = 750 #das ist PI-model abhängig und kmeans abhängig
    hist_peak = np.array(hist_peak)
    upper_peak = np.array(upper_peak)
    mask11 = hist_peak >= schwelle      # Ground Truth: Überschwemmung
    mask12 = hist_peak < schwelle       # Ground Truth: keine Überschwemmung

    mask21 = upper_peak >= schwelle     # Prediction: Warnung
    mask22 = upper_peak < schwelle      # Prediction: keine Warnung

        # TP, FN, FP, TN
    TP = np.sum(mask11 & mask21)
    FN = np.sum(mask11 & mask22)
    FP = np.sum(mask12 & mask21)
    TN = np.sum(mask12 & mask22)

    #print(f"Warn.und HW.: {TP}, Entwarn. aber HW: {FN}, Warn. aber K.HW: {FP}, Entwarn. und K. HW: {TN} ")
    return int(TP), int(FN), int(FP), int(TN)
def inference_live(array, start):
    #temp_arr = []
    #historic_prediction = []
    #historic_prediction_temp = []
    results = []
    model, kmeans, interval_matrix, chunk, cluster, time_teile, time_labels_klein = get_models()
    for i in range(1):
        historic_datum = array[ i * 60, :12].reshape(12, )
        historic_stand = array[ i * 60 : ( i * 60) + 840, 12].reshape(840, )
        historic_vergleich = array[( i * 60) + 840 : ( i * 60) + 840 + 720, 12]*1000
        temp_arr = np.concatenate((historic_datum, historic_stand))
        blaue_kurve = historic_stand
        historic_prediction = []
        for j in range(2):
            historic_prediction_temp = model.predict(temp_arr.reshape(1, -1), verbose=0).reshape(360, )
            historic_prediction = np.append(historic_prediction, historic_prediction_temp)
            temp_arr = temp_arr[12:]
            temp_arr = np.concatenate((temp_arr, historic_prediction_temp))
            temp_arr = temp_arr[360:]
            temp_arr = np.concatenate((array[ i * 60 + (j + 1) * 360, :12].reshape(12, ), temp_arr))
        historic_prediction = historic_prediction * 1000
        cluster_labels2 = assign_clusters_inference( historic_prediction, chunk, kmeans)  # chunk, kmeans global
        q_low_hist, q_up_hist = get_quantile_bounds_from_labels(cluster_labels2, time_labels_klein, interval_matrix)
        lower_hist = historic_prediction - np.abs(q_low_hist)
        upper_hist = historic_prediction + np.abs(q_up_hist)
        zeit2 = stunden_danach(start, 24)
        zeit1 = stunden_zurueck(start, 2)
        zeile = [zeit1, zeit2, np.max(np.array(historic_vergleich)), np.max(np.array(upper_hist)), np.array(historic_prediction), np.array(blaue_kurve)*1000, np.array(lower_hist), np.array(upper_hist)]
        results.append(zeile)
    return results
def inference_live2( array, start):
    #temp_arr = []
    #historic_prediction = []
    #historic_prediction_temp = []
    results = []
    steps = len(array)//720 -1
    model, kmeans, interval_matrix, chunk, cluster, time_teile, time_labels_klein = get_models()
    for i in range(steps):
        historic_datum = array[ i * 720, :12].reshape(12, )
        historic_stand = array[ i * 720 : ( i * 720) + 840, 12].reshape(840, )
        historic_vergleich = array[( i * 720) + 840 : ( i * 720) + 840 + 720, 12]*1000
        temp_arr = np.concatenate((historic_datum, historic_stand))
        historic_prediction = []
        for j in range(2):
            historic_prediction_temp = model.predict(temp_arr.reshape(1, -1), verbose=0).reshape(360, )
            historic_prediction = np.append(historic_prediction, historic_prediction_temp)
            temp_arr = temp_arr[12:]
            temp_arr = np.concatenate((temp_arr, historic_prediction_temp))
            temp_arr = temp_arr[360:]
            temp_arr = np.concatenate((array[ i * 720 + (j + 1) * 360, :12].reshape(12, ), temp_arr))
        historic_prediction = historic_prediction * 1000
        cluster_labels2 = assign_clusters_inference( historic_prediction, chunk, kmeans)  # chunk, kmeans global
        q_low_hist, q_up_hist = get_quantile_bounds_from_labels(cluster_labels2, time_labels_klein, interval_matrix)
        lower_hist = historic_prediction - np.abs(q_low_hist)
        upper_hist = historic_prediction + np.abs(q_up_hist)
        zeit1 = start
        zeit2 = stunden_danach(start, 12)
        start = stunden_danach(start, 1)
        zeile = [zeit1, zeit2, np.max(np.array(historic_vergleich)), np.max(np.array(upper_hist)), np.array(historic_prediction), np.array(historic_vergleich), np.array(lower_hist), np.array(upper_hist)]
        results.append(zeile)
    return results

def extract_and_transform_live(stunden, inference_func):
    json_daten = osw_api_extract(stunden_zurueck(now_berlin_time(), stunden), now_berlin_time(), api_url_template)
    df3 = transform(json_daten)
    results = inference_func( df3, stunden_zurueck(now_berlin_time(), stunden))
    return results
#########################################################################
def load_time_series(conn_str, startzeit_iso, endzeit_iso):
    with psy.connect(conn_str) as conn:
        with conn.cursor() as cur:
            current_start = startzeit_iso
            results = []

            while True:
                # Hole den ersten Datensatz, dessen startzeit = current_start ist
                cur.execute("""
                    SELECT id, startzeit, endzeit
                    FROM zeitreihe_metadata
                    WHERE startzeit = %s
                      AND endzeit <= %s
                    ORDER BY endzeit ASC
                    LIMIT 1
                """, (current_start, endzeit_iso))
                row = cur.fetchone()

                if not row:
                    print(f"⚠️ Keine weiteren Zeiträume gefunden ab {current_start}. Abbruch.")
                    break

                metadata_id, startzeit, endzeit = row

                # Hole Zeitreihendaten für dieses metadata_id
                cur.execute("""
                    SELECT minute_value, prediction_value, historic_value, lowerpi_value, upperpi_value
                    FROM zeitreihe_daten
                    WHERE metadata_id = %s
                    ORDER BY minute_value ASC
                """, (metadata_id,))
                data_rows = cur.fetchall()

                # Check, ob wir weniger als 720 Minuten haben
                num_points = len(data_rows)
                if num_points < 144:
                    print(f"⚠️ Zeitreihe bei {startzeit} - {endzeit} hat nur {num_points} Punkte (truncated).")
                elif num_points > 144:
                    print(f"⚠️ Zeitreihe länger als 720 Punkte, wird abgeschnitten.")
                    data_rows = data_rows[:144]

                # Speichere Ergebnis
                results.append({
                    "metadata_id": metadata_id,
                    "startzeit": startzeit,
                    "endzeit": endzeit,
                    "data": data_rows
                })

                # Nächster Startzeitpunkt = aktueller Endzeitpunkt
                current_start = endzeit.strftime("%Y-%m-%dT%H:%M:%S")

                # Prüfen, ob wir über Endzeit hinaus sind
                if current_start >= endzeit_iso:
                    print("✅ Endzeit erreicht oder überschritten. Fertig.")
                    break

            return results
def stretch_array(arr):
    arr = np.array(arr)
    n = len(arr)

    if n < 2:
        raise ValueError("Array muss mindestens 2 Elemente haben, um zu interpolieren.")

    # Für jedes Paar von Punkten wollen wir 5 Punkte insgesamt (Originalpunkte + 4 dazwischen)
    # Z.B. [a, b] -> [a, ..., b] mit 5 Punkten zwischen a und b
    # Wir bauen einen Index-Vektor von 0 bis n-1, jeweils in 1er-Schritten
    old_indices = np.arange(n)

    # Neuer Index-Vektor mit feinerer Auflösung
    new_indices = np.linspace(0, n - 1, (n - 1) * 5 + 1)

    # Interpolation
    new_arr = np.interp(new_indices, old_indices, arr)

    return new_arr.tolist()
def extract_and_stretch(results):
    """Extrahiert Spalten aus einem Block und interpoliert sie."""
    pred_ = []
    hist_ = []
    lower_ = []
    upper_ = []
    for i in range(len(results)):
        pred  = stretch_array([row[1] for row in results[i]['data']])
        hist  = stretch_array([row[2] for row in results[i]['data']])
        lower = stretch_array([row[3] for row in results[i]['data']])
        upper = stretch_array([row[4] for row in results[i]['data']])
        pred_.append(pred)
        hist_.append(hist)
        lower_.append(lower)
        upper_.append(upper)
    return np.array(pred_).flatten(), np.array(hist_).flatten(), np.array(lower_).flatten(), np.array(upper_).flatten()
def plot_time_series(pred, hist, lower, upper):
    """Plottet die Zeitreihen."""

    plt.figure(figsize=(14, 6))
    plt.plot(pred, label="Prediction")
    plt.plot(hist, label="Historic")
    plt.plot(lower, label="Lower PI", linestyle="--")
    plt.plot(upper, label="Upper PI", linestyle="--")
    plt.fill_between(range(len(pred)), lower, upper, alpha=0.2, color='gray')
    plt.legend()
    plt.title("Interpolierte Zeitreihen")
    plt.grid(True)
    plt.show()
def plot_time_series(pred, hist, lower, upper, save_path=None):
    plt.figure(figsize=(14, 6))
    plt.plot(pred, label="Prediction")
    plt.plot(hist, label="Historic")
    plt.plot(lower, label="Lower PI", linestyle="--")
    plt.plot(upper, label="Upper PI", linestyle="--")
    plt.fill_between(range(len(pred)), lower, upper, alpha=0.2, color='gray')
    plt.legend()
    plt.title("Interpolierte Zeitreihen")
    plt.grid(True)
    if save_path:
        plt.savefig(save_path)
        print(f"📈 Plot gespeichert unter {save_path}")
    plt.show()
def save_to_excel(pred, hist, lower, upper, filename):
    df = pd.DataFrame({
        "Prediction": pred,
        "Historic": hist,
        "Lower_PI": lower,
        "Upper_PI": upper
    })
    df.index.name = "Index"
    df.to_excel(filename)
    print(f"📄 Daten gespeichert in {filename}")
def export_to_excel_with_chart(pred, hist, lower, upper, filename="zeitreihe.xlsx"):
    # Baue DataFrame
    df = pd.DataFrame({
        "Prediction": pred,
        "Historic": hist,
        "Lower_PI": lower,
        "Upper_PI": upper
    })

    # Speichere erstmal ohne Chart
    df.to_excel(filename, index=False)

    # Lade Workbook mit openpyxl
    wb = load_workbook(filename)
    ws = wb.active

    # Erstelle LineChart
    chart = LineChart()
    chart.title = "Zeitreihe mit Prediction, Historic & PI"
    chart.y_axis.title = "Wert"
    chart.x_axis.title = "Zeitpunkte"

    # Referenz: alle Spalten (A bis D), Zeilen 2 bis letzte
    data = Reference(ws, min_col=1, max_col=4, min_row=1, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)

    # Positioniere Chart (z. B. ab Zelle F2)
    ws.add_chart(chart, "F2")

    # Speichern
    wb.save(filename)
    print(f"✅ Excel-Datei mit Diagramm gespeichert: {filename}")
    return filename


app = FastAPI()

@app.get("/download-excel")
def download_excel2():
    conn_str = "postgresql://neondb_owner:npg_mPqZi9CG2txF@ep-divine-mud-a90zxdvg-pooler.gwc.azure.neon.tech/neondb?sslmode=require&channel_binding=require"


    results = np.array(load_time_series(conn_str, start_iso, end_iso))
    pred, hist, lower, upper = extract_and_stretch(results)

    # Statt Dateipfad -> BytesIO
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Beispiel: deine Daten in Excel schreiben
        pd.DataFrame(pred).to_excel(writer, sheet_name='Predictions')
        pd.DataFrame(hist).to_excel(writer, sheet_name='History')
        # du kannst hier beliebig Tabellen hinzufügen

    output.seek(0)  # ganz wichtig!

    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment; filename=report.xlsx"}
    )


class TimeRange(BaseModel):
    start_iso: str
    end_iso: str

@app.post("/download-excel")
def download_excel(time_range: TimeRange):

    results = np.array(load_time_series(conn_str, time_range.start_iso, time_range.end_iso))
    pred, hist, lower, upper = extract_and_stretch(results)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(pred).to_excel(writer, sheet_name='Predictions')
        pd.DataFrame(hist).to_excel(writer, sheet_name='History')
        # weitere Tabellen hinzufügen

    output.seek(0)

    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment; filename=report.xlsx"}
    )

@app.get("/get-live")
def live():
    results = extract_and_transform_live(12,inference_live)
    result_json_ready = []
    for zeile in results:
        result_json_ready.append({
            "zeit1": str(zeile[0]),
            "zeit2": str(zeile[1]),
            "max_vergleich": zeile[2],
            "max_upper": zeile[3],
            "historic_prediction": zeile[4].tolist(),
            "blaue_kurve": zeile[5].tolist(),
            "lower_hist": zeile[6].tolist(),
            "upper_hist": zeile[7].tolist(),
        })
    return JSONResponse(content={"results": result_json_ready})

@app.get("/get-live2")
def live2():
    results = extract_and_transform_live(120, inference_live2)
    # Neue leere Listen für jede "Spalte"
    grouped = {
        "zeit1": [],
        "zeit2": [],
        "max_vergleich": [],
        "max_upper": [],
        "historic_prediction": [],
        "historic_vergleich": [],
        "lower_hist": [],
        "upper_hist": [],
        "statistics": []
    }
    for zeile in results:
        grouped["zeit1"].append(str(zeile[0]))
        grouped["zeit2"].append(str(zeile[1]))
        grouped["max_vergleich"].append(zeile[2])
        grouped["max_upper"].append(zeile[3])
        grouped["historic_prediction"].append(zeile[4].tolist())
        grouped["historic_vergleich"].append(zeile[5].tolist())
        grouped["lower_hist"].append(zeile[6].tolist())
        grouped["upper_hist"].append(zeile[7].tolist())

    TP, FN, FP, TN = warnung_generator(np.array(grouped["max_upper"]).flatten(), np.array(grouped["max_vergleich"]).flatten())
    grouped["statistics"] = [TP, FN, FP, TN]

    return JSONResponse(content=grouped)



if __name__ == "__main__":
    #uvicorn.run("db_2_excel:app", host="127.0.0.1", port=8000, reload=True)
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run("db_2_excel:app", host="0.0.0.0", port=port, reload=False)
